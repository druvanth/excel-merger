import tempfile
from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook, Workbook
from tempfile import NamedTemporaryFile
from copy import copy as style_copy
from openpyxl.worksheet.table import Table

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'files[]' not in request.files:
        return 'No file part in the request'

    files = request.files.getlist('files[]')
    if len(files) == 0:
        return 'No files selected'

    # Create a temporary file to store the merged workbook
    temp_file = NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_file.close()

    # Merge the uploaded files
    merge_excel_files(files, temp_file.name)

    # Serve the merged file to the user
    return send_file(temp_file.name, as_attachment=True, download_name='merged_workbook.xlsx')

def merge_excel_files(files, output_path):
    merged_workbook = Workbook()

    for file in files:
        # Load the source workbook
        source_workbook = load_workbook(file)

        # Iterate over the sheets in the source workbook and copy them to the merged workbook
        for source_sheet_name in source_workbook.sheetnames:
            source_sheet = source_workbook[source_sheet_name]

            if source_sheet_name in merged_workbook.sheetnames:
                target_sheet = merged_workbook[source_sheet_name]
            else:
                target_sheet = merged_workbook.create_sheet(title=source_sheet_name)

            # Copy cell values
            for row in source_sheet.iter_rows(values_only=True):
                target_sheet.append(row)

            # Copy sheet formatting
            for row in source_sheet.iter_rows(min_row=1):
                for cell in row:
                    if cell.value is not None:
                        target_sheet[cell.coordinate].font = style_copy(cell.font)
                        # Copy fill style only if cell has a fill
                        if cell.fill and (cell.fill.fill_type == "solid" ):
                            target_sheet[cell.coordinate].fill = style_copy(cell.fill)   
                        target_sheet[cell.coordinate].border = style_copy(cell.border)
                        target_sheet[cell.coordinate].alignment = style_copy(cell.alignment)
                        target_sheet[cell.coordinate].number_format = style_copy(cell.number_format)
                        target_sheet[cell.coordinate].protection = style_copy(cell.protection)
                        target_cell = target_sheet[cell.coordinate]

            # Copy merged cells
            merged_ranges = source_sheet.merged_cells.ranges
            if merged_ranges:
                for merged_range in merged_ranges:
                    target_sheet.merge_cells(str(merged_range))         

            # Copy images
            if source_sheet._images:
                for image in source_sheet._images:
                    target_sheet.add_image(image)

            # Copy table formatting if tables exist
            if source_sheet.tables:
                for source_table in source_sheet.tables.values():
                    table_range = source_table.ref
                    target_table_name = f"{source_sheet_name}_{source_table.name}" # Changing the table names so that no 2 tables have same names.
                    target_table = Table(displayName=target_table_name.name, ref=table_range)
                        
                    # Copy table formatting and cell values
                    for row in source_sheet[table_range]:
                        for cell in row:
                            target_cell = target_sheet[cell.coordinate]
                            target_cell.value = cell.value

                            # Copy cell formatting
                            if cell.has_style:
                                source_style = cell.style
                                target_sheet[cell.coordinate].font = style_copy(cell.font)
                                # Copy fill style only if cell has a fill
                                if cell.fill and (cell.fill.fill_type == "solid" ):
                                    target_sheet[cell.coordinate].fill = style_copy(cell.fill)

                                target_sheet[cell.coordinate].border = style_copy(cell.border)
                                target_sheet[cell.coordinate].alignment = style_copy(cell.alignment)
                                target_sheet[cell.coordinate].number_format = style_copy(cell.number_format)
                                target_sheet[cell.coordinate].protection = style_copy(cell.protection)

                    target_sheet.add_table(target_table)

            # Preserve filters
            if source_sheet.auto_filter:
                target_sheet.auto_filter.ref = source_sheet.auto_filter.ref

            # Preserve hidden rows
            for row in source_sheet.row_dimensions:
                if source_sheet.row_dimensions[row].hidden:
                    target_sheet.row_dimensions[row].hidden = True

            # Preserve hidden columns
            for column_letter, column_dimension in source_sheet.column_dimensions.items():
                if column_dimension.hidden:
                    target_column_dimensions = target_sheet.column_dimensions
                    if column_letter not in target_column_dimensions:
                        target_column_dimensions[column_letter] = column_dimension
                    else:
                        target_column_dimensions[column_letter].hidden = True

            # Preserve column width
            for column_letter, column_dimension in source_sheet.column_dimensions.items():
                if column_dimension.width is not None:
                    target_column_dimensions = target_sheet.column_dimensions
                    if column_letter not in target_column_dimensions:
                        target_column_dimensions[column_letter] = column_dimension
                    else:
                        target_column_dimensions[column_letter].width = column_dimension.width

            # Preserve row height
            for row, row_dimension in source_sheet.row_dimensions.items():
                if row_dimension.height is not None:
                    target_row_dimensions = target_sheet.row_dimensions
                    if row not in target_row_dimensions:
                        target_row_dimensions[row] = row_dimension
                    else:
                        target_row_dimensions[row].height = row_dimension.height

        print(file)

    # Remove default sheet created by Workbook()
    default_sheet = merged_workbook['Sheet']
    merged_workbook.remove(default_sheet)

    # Save the merged workbook to the specified output path
    merged_workbook.save(filename=output_path)

if __name__ == '__main__':
    app.run()
